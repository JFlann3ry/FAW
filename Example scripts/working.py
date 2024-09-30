from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('test.xlsx')

# Select the active sheet
ws = wb['xRecipeTags']

# Define the RecipeID and the list of TagIDs
recipe_id = 69
tag_ids = [1, 2, 3]

# Find the last row in the ID column
last_row = ws.max_row

# Increment the last ID by 1
last_id = ws.cell(row=last_row, column=1).value if last_row > 1 else 0
new_id = last_id + 1

# Append a new row for each TagID
for tag_id in tag_ids:
    new_row = [new_id, recipe_id, tag_id]
    ws.append(new_row)
    new_id += 1

# Save the workbook
wb.save('test.xlsx')
