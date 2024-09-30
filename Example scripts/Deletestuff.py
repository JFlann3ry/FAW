from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('test.xlsx')

# Select the sheet
ws = wb['xRecipeTags']

# Iterate through the rows backwards to safely delete rows
for row in range(ws.max_row, 1, -1):
    if ws.cell(row=row, column=2).value == 69:  # Assuming RecipeID is in the 2nd column
        ws.delete_rows(row)

# Save the workbook
wb.save('test.xlsx')
